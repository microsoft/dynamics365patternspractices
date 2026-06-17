from __future__ import annotations

import csv
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def find_source_files(source_dir: str) -> list[Path]:
    root = Path(source_dir)
    files = []
    for pattern in ("*.xlsx", "*.xlsm", "*.csv", "*.tsv"):
        files.extend(root.glob(pattern))
    return sorted(
        f
        for f in files
        if not f.name.startswith("~$")
        and "template" not in f.name.lower()
        and "guideline" not in f.name.lower()
    )


def load_rows(path: Path) -> list[dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix in {".csv", ".tsv"}:
        return _load_delimited(path, "\t" if suffix == ".tsv" else ",")
    if suffix in {".xlsx", ".xlsm"}:
        return _load_workbook(path)
    raise ValueError(f"Unsupported source file: {path}")


def _load_delimited(path: Path, delimiter: str) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle, delimiter=delimiter)
        rows = []
        for row_number, row in enumerate(reader, start=2):
            if any((v or "").strip() for v in row.values()):
                clean = {str(k or "").strip(): v for k, v in row.items() if k}
                clean["_source_file"] = path.name
                clean["_source_sheet"] = ""
                clean["_source_row"] = row_number
                rows.append(clean)
        return rows


def _load_workbook(path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(path, data_only=True, read_only=True)
    rows: list[dict[str, Any]] = []
    for sheet in wb.worksheets:
        extracted = _load_sheet(path, sheet)
        if extracted:
            rows.extend(extracted)
    return rows


def _load_sheet(path: Path, sheet) -> list[dict[str, Any]]:
    raw = list(sheet.iter_rows(values_only=True))
    header_index = None
    headers: list[str] = []
    for idx, values in enumerate(raw):
        candidate = [str(v or "").strip() for v in values]
        non_empty = [v for v in candidate if v]
        if len(non_empty) >= 2:
            header_index = idx
            headers = candidate
            break
    if header_index is None:
        return []

    rows = []
    for offset, values in enumerate(raw[header_index + 1 :], start=header_index + 2):
        if _looks_like_secondary_header(headers, values):
            continue
        row = {headers[i]: values[i] if i < len(values) else None for i in range(len(headers)) if headers[i]}
        if any(v not in (None, "") for v in row.values()):
            row["_source_file"] = path.name
            row["_source_sheet"] = sheet.title
            row["_source_row"] = offset
            rows.append(row)
    return rows


def _looks_like_secondary_header(headers: list[str], values: tuple[Any, ...]) -> bool:
    matches = 0
    checked = 0
    for i, header in enumerate(headers):
        if not header or i >= len(values) or values[i] in (None, ""):
            continue
        checked += 1
        if _header_base(header) == _header_base(str(values[i])):
            matches += 1
    return checked >= 4 and matches / checked >= 0.6


def _header_base(value: str) -> str:
    value = re.sub(r"\s+ms\s+bpc$", "", value.strip(), flags=re.IGNORECASE)
    return " ".join(value.lower().split())





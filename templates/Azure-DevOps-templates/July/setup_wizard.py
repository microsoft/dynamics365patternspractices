import argparse
import os
import subprocess
import sys
from dataclasses import dataclass
from typing import Dict, List, Optional

from ado_setup_config import AdoSetupConfig, load_config


SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))


@dataclass(frozen=True)
class Phase:
    number: int
    name: str
    script: str
    required_sheets: List[str]


PHASES = [
    Phase(
        1,
        "Create process, project, work item types, fields, and picklists",
        "1_ADO_Creation_Script.py",
        ["Work item types", "Fields", "Picklists"],
    ),
    Phase(
        2,
        "Create work item page layouts",
        "2_ADO_Page_Layout_Script_Threaded.py",
        ["Work item types", "Fields"],
    ),
    Phase(
        3,
        "Create teams, area paths, and team area assignments",
        "3_ADO_Teams_Areas_Script.py",
        ["Area paths"],
    ),
    Phase(
        4,
        "Configure backlogs, iterations, and team settings",
        "4_ADO_Backlog_Config_Script.py",
        ["Backlogs", "Work item types", "Iteration paths", "Teams", "Area paths"],
    ),
    Phase(
        5,
        "Import Business Process Catalog work items",
        "5_BPC_Catalog_Import.py",
        [],
    ),
    Phase(
        6,
        "Generate HTML setup summary report",
        "6_Generate_HTML_Report.py",
        [],
    ),
]


def _build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description=(
            "Guided setup wizard for the Microsoft Business Process Catalog Azure DevOps template. "
            "Common Azure DevOps values can be passed with --ado-org-url, --ado-project, "
            "--process-name, --excel-file, or environment variables."
        )
    )
    parser.add_argument("--start-at", type=int, choices=[1, 2, 3, 4, 5, 6], default=1, help="First phase to run.")
    parser.add_argument("--stop-after", type=int, choices=[1, 2, 3, 4, 5, 6], default=6, help="Last phase to run.")
    parser.add_argument("--catalog-source-dir", help="Folder containing the four BPC source files. Defaults to the parent folder of a 'Python Scripts' template folder.")
    parser.add_argument("--catalog-output", help="Base output folder for phase 5 import logs and ID maps.")
    parser.add_argument("--catalog-parallel-workers", default="4", help="Parallel worker count for phase 5 catalog import.")
    parser.add_argument(
        "--skip-excel-validation",
        action="store_true",
        help="Skip the early Excel workbook validation check.",
    )
    parser.add_argument(
        "--skip-ado-validation",
        action="store_true",
        help="Skip the early Azure DevOps PAT/org validation check.",
    )
    parser.add_argument(
        "--skip-result-validation",
        action="store_true",
        help="Skip validation after each completed phase.",
    )
    return parser


def _print_header(title: str) -> None:
    print()
    print("=" * len(title))
    print(title)
    print("=" * len(title))


def _load_workbook_sheet_names(excel_file: str) -> List[str]:
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError(
            "The openpyxl package is required to validate the Excel template. "
            "Install it with: python -m pip install openpyxl"
        ) from exc

    try:
        workbook = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
    except Exception as exc:
        raise RuntimeError(
            f"Could not open the Excel template '{excel_file}'. Confirm it is a valid .xlsx file and is closed in Excel."
        ) from exc

    return list(workbook.sheetnames)


def validate_excel(config: AdoSetupConfig, selected_phases: List[Phase]) -> None:
    _print_header("Validating Excel template")
    if not os.path.exists(config.excel_file):
        raise FileNotFoundError(f"Excel template not found: {config.excel_file}")

    sheet_names = set(_load_workbook_sheet_names(config.excel_file))
    required = []
    for phase in selected_phases:
        required.extend(phase.required_sheets)

    missing = sorted({sheet for sheet in required if sheet not in sheet_names})
    if missing:
        raise RuntimeError("The Excel template is missing required sheets: " + ", ".join(missing))

    print("Excel template found and required sheets are present.")


def validate_ado_access(config: AdoSetupConfig) -> None:
    _print_header("Validating Azure DevOps access")
    try:
        import requests
    except ImportError as exc:
        raise RuntimeError("The requests package is required. Install it with: python -m pip install requests") from exc

    response = requests.get(
        f"{config.ado_org_url}/_apis/projects?api-version=7.1",
        auth=("", config.pat),
        timeout=30,
    )
    if response.status_code in (401, 403):
        raise RuntimeError(
            "Azure DevOps rejected the PAT or the account does not have enough access. "
            "Confirm the PAT scopes and that you are a Project Collection Administrator."
        )
    if response.status_code >= 400:
        raise RuntimeError(f"Could not validate Azure DevOps access: {response.status_code} - {response.text}")
    print("Azure DevOps organization and PAT were accepted.")


def _default_catalog_source(excel_file: str) -> str:
    excel_dir = os.path.dirname(os.path.abspath(excel_file))
    if os.path.basename(excel_dir).lower() == "python scripts":
        return os.path.dirname(excel_dir)
    return excel_dir


def _phase_env(config: AdoSetupConfig, args: argparse.Namespace) -> Dict[str, str]:
    env = os.environ.copy()
    env.update(
        {
            "BPC_ADO_ORG_URL": config.ado_org_url,
            "BPC_ADO_PROJECT": config.ado_project,
            "BPC_ADO_PROCESS_NAME": config.process_name,
            "BPC_ADO_PAT": config.pat,
            "BPC_ADO_EXCEL_FILE": config.excel_file,
            "BPC_ADO_CATALOG_SOURCE_DIR": args.catalog_source_dir or _default_catalog_source(config.excel_file),
            "BPC_ADO_IMPORT_OUTPUT": args.catalog_output or os.path.join(SCRIPT_DIR, "out"),
            "BPC_ADO_IMPORT_PARALLEL_WORKERS": str(args.catalog_parallel_workers),
        }
    )
    return env


def run_phase(phase: Phase, config: AdoSetupConfig, args: argparse.Namespace) -> None:
    _print_header(f"Phase {phase.number}: {phase.name}")
    script_path = os.path.join(SCRIPT_DIR, phase.script)
    if not os.path.exists(script_path):
        raise FileNotFoundError(f"Required script not found: {script_path}")

    result = subprocess.run(
        [sys.executable, script_path, "--non-interactive"],
        cwd=SCRIPT_DIR,
        env=_phase_env(config, args),
        check=False,
    )
    if result.returncode != 0:
        raise RuntimeError(f"Phase {phase.number} failed. Review the phase log file before rerunning the wizard.")


def selected_phases(start_at: int, stop_after: int) -> List[Phase]:
    if start_at > stop_after:
        raise ValueError("--start-at must be less than or equal to --stop-after.")
    return [phase for phase in PHASES if start_at <= phase.number <= stop_after]


def main() -> int:
    args, common_args = _build_parser().parse_known_args()
    config = load_config(
        default_log_file="bpc_ado_setup_wizard.log",
        require_process=True,
        argv=common_args,
        ignore_unknown_args=False,
    )
    phases = selected_phases(args.start_at, args.stop_after)

    print()
    print("This wizard keeps your PAT in memory for this run only. It does not write the PAT to script files.")

    if not args.skip_excel_validation:
        validate_excel(config, phases)
    if any(phase.number == 5 for phase in phases):
        catalog_source = args.catalog_source_dir or _default_catalog_source(config.excel_file)
        if not os.path.isdir(catalog_source):
            raise FileNotFoundError(f"Catalog source folder not found for phase 5: {catalog_source}")
    if not args.skip_ado_validation:
        validate_ado_access(config)

    for phase in phases:
        run_phase(phase, config, args)
        if not args.skip_result_validation and phase.number <= 4:
            from ado_setup_validation import validate_phase

            validate_phase(phase.number, config)

    _print_header("Setup complete")
    print("All selected phases finished. Review the log files, import output, and Azure DevOps project.")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except KeyboardInterrupt:
        print("\nSetup canceled by user.")
        raise SystemExit(130)
    except Exception as exc:
        print(f"\nERROR: {exc}")
        raise SystemExit(1)

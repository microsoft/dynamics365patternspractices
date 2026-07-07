from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


SYSTEM_FIELD_ALIASES = {
    "title": "System.Title",
    "name": "System.Title",
    "work item title": "System.Title",
    "description": "System.Description",
    "acceptance criteria": "Microsoft.VSTS.Common.AcceptanceCriteria",
    "priority": "Microsoft.VSTS.Common.Priority",
    "state": "System.State",
    "reason": "System.Reason",
    "area path": "System.AreaPath",
    "iteration path": "System.IterationPath",
    "assigned to": "System.AssignedTo",
    "tags": "System.Tags",
}

BUILT_IN_FIELD_ALIASES = {
    "process sequence id": "MSBPC.processsequenceid",
    "alternate process sequence id": "MSBPC.alternateprocesssequenceid",
    "microsoft id": "MSBPC.microsoftid",
    "partner id": "MSBPC.partnerid",
    "mavim id": "MSBPC.mavimid",
    "catalog status": "MSBPC.catalogstatus",
    "article status": "MSBPC.articlestatus",
    "business process flow status": "MSBPC.businessprocessflowstatus",
    "scope": "MSBPC.scope",
    "fit gap status": "MSBPC.fitgapstatus",
    "gap solution approach": "MSBPC.gapsolutionapproach",
    "business process owner": "MSBPC.businessprocessowner",
    "workstream lead": "MSBPC.workstreamlead",
    "subject matter expert": "MSBPC.subjectmatterexpert",
    "workstream": "MSBPC.workstream",
    "application family": "MSBPC.applicationfamily",
    "products": "MSBPC.products",
    "industries": "MSBPC.industries",
    "module": "MSBPC.module",
    "menu path": "MSBPC.menupath",
    "menu item name": "MSBPC.menuitemname",
    "success by design phase": "MSBPC.successbydesignphase",
    "microsoft references": "MSBPC.microsoftreferences",
    "partner references": "MSBPC.partnerreferences",
    "internal references": "MSBPC.internalreferences",
    "update comments": "MSBPC.updatecomments",
    "steps": "Microsoft.VSTS.TCM.Steps",
    "automation status": "Microsoft.VSTS.TCM.AutomationStatus",
}

FIELD_LABELS = {v.lower(): k for k, v in (SYSTEM_FIELD_ALIASES | BUILT_IN_FIELD_ALIASES).items()}


@dataclass
class FieldDef:
    label: str
    reference_name: str
    field_type: str = ""
    required: str = ""
    default_value: str = ""
    multi_select: bool = False


@dataclass
class TemplateMapping:
    fields_by_label: dict[str, FieldDef] = field(default_factory=dict)
    fields_by_ref: dict[str, FieldDef] = field(default_factory=dict)
    aliases: dict[str, str] = field(default_factory=lambda: SYSTEM_FIELD_ALIASES | BUILT_IN_FIELD_ALIASES)
    applicable_fields_by_wit: dict[str, set[str]] = field(default_factory=dict)
    work_item_types: set[str] = field(default_factory=set)

    def field_ref_for_header(self, header: str) -> str | None:
        key = normalize(header)
        if key in self.aliases:
            return self.aliases[key]
        if key.startswith("msbpc."):
            return header.strip()
        if key.startswith("microsoft.") or key.startswith("system."):
            return header.strip()
        return None

    def field_def(self, reference_name: str) -> FieldDef | None:
        return self.fields_by_ref.get(reference_name.lower())


def normalize(value: Any) -> str:
    return " ".join(str(value or "").strip().lower().replace("_", " ").split())


def load_template(path: str | None) -> TemplateMapping:
    mapping = TemplateMapping()
    for label, ref in mapping.aliases.items():
        fd = FieldDef(label=label.title(), reference_name=ref)
        mapping.fields_by_label[normalize(label)] = fd
        mapping.fields_by_ref[ref.lower()] = fd
    if not path:
        return mapping

    wb = load_workbook(path, data_only=True, read_only=True)
    if "Fields" in wb.sheetnames:
        _load_fields_sheet(wb["Fields"], mapping)
    if "Work item types" in wb.sheetnames:
        _load_work_item_types_sheet(wb["Work item types"], mapping)
    return mapping


def _rows(sheet) -> tuple[list[str], list[dict[str, Any]]]:
    iterator = sheet.iter_rows(values_only=True)
    headers = [str(c or "").strip() for c in next(iterator)]
    rows = []
    for values in iterator:
        row = {headers[i]: values[i] if i < len(values) else None for i in range(len(headers))}
        if any(v not in (None, "") for v in row.values()):
            rows.append(row)
    return headers, rows


def _load_fields_sheet(sheet, mapping: TemplateMapping) -> None:
    _, rows = _rows(sheet)
    for row in rows:
        label = row.get("Label") or row.get("Field name")
        ref = row.get("Reference name")
        if not label or not ref:
            continue
        fd = FieldDef(
            label=str(label).strip(),
            reference_name=str(ref).strip(),
            field_type=str(row.get("Field type") or ""),
            required=str(row.get("Required") or ""),
            default_value=str(row.get("Default value") or "").strip(),
            multi_select=normalize(row.get("Multi-select field")) == "yes",
        )
        for alias in {fd.label, row.get("Field name"), ref}:
            if alias:
                mapping.aliases[normalize(alias)] = fd.reference_name
        mapping.fields_by_label[normalize(fd.label)] = fd
        mapping.fields_by_ref[fd.reference_name.lower()] = fd


def _load_work_item_types_sheet(sheet, mapping: TemplateMapping) -> None:
    headers, rows = _rows(sheet)
    metadata = {
        "index",
        "Work item type",
        "Custom work item type",
        "Inherit from",
        "Reference name",
        "Color",
        "Icon",
        "Help text",
        "Purpose",
        "Backlog type",
        "Backlog name",
    }
    field_headers = [h for h in headers if h and h not in metadata]
    for row in rows:
        wit = row.get("Work item type")
        if not wit or normalize(row.get("Custom work item type")) == "disabled":
            continue
        wit = str(wit).strip()
        mapping.work_item_types.add(wit)
        refs = set()
        for header in field_headers:
            marker = normalize(row.get(header))
            if marker in {"x", "h"}:
                ref = mapping.field_ref_for_header(header)
                if ref:
                    refs.add(ref)
        mapping.applicable_fields_by_wit[wit.lower()] = refs
